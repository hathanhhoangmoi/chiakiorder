import httpx
import io
import json
import re
from datetime import datetime, timedelta
from sqlalchemy.orm import Session
from models import Order, ShopMeta
from shops_config import SELLER_ID, SELLER_TOKEN
from openpyxl import load_workbook


async def fetch_shop_name(shop_url: str) -> str:
    """Giữ lại cho /api/test-shopname"""
    try:
        async with httpx.AsyncClient(
            timeout=10,
            follow_redirects=True,
            headers={"User-Agent": "Mozilla/5.0"}
        ) as client:
            res = await client.get(shop_url)
            if not res.is_success:
                return shop_url
        patterns = [
            r']*class=["\']store-title["\'][^>]*>(.*?)',
            r'class=["\']store-title["\'][^>]*>(.*?)<',
            r'store-title["\']>(.*?)<',
        ]
        for pattern in patterns:
            m = re.search(pattern, res.text, re.IGNORECASE | re.DOTALL)
            if m:
                name = m.group(1).strip()
                if name:
                    return name
        return shop_url
    except Exception as e:
        print(f"[fetch_name] Error {shop_url}: {e}")
        return shop_url


def parse_excel(content: bytes, shop_id: str, shop_name: str) -> list[dict] | None:
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return None

        headers = [str(c).strip() if c else "" for c in rows[0]]
        print(f"[parse] headers: {headers}")

        def find_col(keywords):
            for kw in keywords:
                for i, h in enumerate(headers):
                    if kw.lower() in h.lower():
                        return i
            return None

        col_code     = find_col(["mã đơn hàng", "mã đơn", "order_id"])
        col_customer = find_col(["người đặt hàng", "người đặt", "đặt hàng", "khách"])
        col_buyer    = find_col(["tên người nhận", "người nhận", "buyer"])
        col_phone    = find_col(["sđt", "điện thoại", "phone", "số điện thoại"])
        col_address  = find_col(["địa chỉ", "address"])
        col_product  = find_col(["tên sản phẩm", "ten san pham", "tên hàng", "product name", "product"])
        col_qty      = find_col(["số lượng", "quantity", "qty", "sl"])
        col_total    = find_col(["tổng tiền", "tổng", "total", "amount"])
        col_status   = find_col(["trạng thái", "status"])
        col_date     = find_col(["thời gian đặt hàng", "thời gian đặt", "thời gian", "ngày đặt", "ngày tạo", "ngày", "date", "time"])

        recognized_cols = [
            col_code, col_customer, col_buyer, col_phone, col_address,
            col_product, col_qty, col_total, col_status, col_date
        ]
        if all(idx is None for idx in recognized_cols):
            return None

        def val(row, idx):
            if idx is None or idx >= len(row): return ""
            v = row[idx]
            return str(v).strip() if v is not None else ""

        orders = []
        for i, row in enumerate(rows[1:]):
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            code = f"{shop_id}_{val(row, col_code)}" if val(row, col_code) else f"{shop_id}_{i}"
            if not code or code == "None": continue
            try:
                qty   = int(float(val(row, col_qty)))   if val(row, col_qty)   else 0
                total = float(val(row, col_total))       if val(row, col_total) else 0.0
            except:
                qty, total = 0, 0.0
            orders.append({
                "order_code":    code,
                "shop_id":       shop_id,
                "shop_name":     shop_name,
                "buyer_name":    val(row, col_buyer),
                "customer_name": val(row, col_customer),
                "phone":         val(row, col_phone),
                "address":       val(row, col_address),
                "product":       val(row, col_product),
                "quantity":      qty,
                "total":         total,
                "status":        val(row, col_status),
                "order_date":    val(row, col_date),
                "raw_data":      json.dumps(
                    dict(zip(headers, [str(c) for c in row])),
                    ensure_ascii=False
                ),
            })
        return orders
    except Exception as e:
        print(f"[parse_excel] Error shop {shop_id}: {e}")
        return None


async def sync_shop(
    shop_id: str,
    shop_url: str,
    shop_name: str,
    db: Session,
    cf_chl_tk: str = "",
    cf_clearance: str = "",
) -> int:
    today = datetime.now()
    since = today - timedelta(days=14)
    def fmt(d): return d.strftime("%d/%m/%Y").replace("/", "%2F")
    range_str = f"{fmt(since)}%20-%20{fmt(today)}"

    url = (
        f"https://api.chiaki.vn/api/{shop_id}/export-excel-order"
        f"?source=seller&page_index=1&page_size=500&status=all"
        f"&range_date={range_str}"
        f"&date_type=created_at&order=create-desc"
        f"&Seller_id={SELLER_ID}&Seller_token={SELLER_TOKEN}"
        f"&__cf_chl_tk={cf_chl_tk}"
    )

    headers = {
        "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
        "accept-language": "vi,en-US;q=0.9,en;q=0.8",
        "cache-control": "max-age=0",
        "referer": url,
        "sec-ch-ua": '"Chromium";v="146", "Not-A.Brand";v="24", "Google Chrome";v="146"',
        "sec-ch-ua-arch": '""',
        "sec-ch-ua-bitness": '"64"',
        "sec-ch-ua-full-version": '"146.0.7680.154"',
        "sec-ch-ua-full-version-list": '"Chromium";v="146.0.7680.154", "Not-A.Brand";v="24.0.0.0", "Google Chrome";v="146.0.7680.154"',
        "sec-ch-ua-mobile": "?1",
        "sec-ch-ua-model": '"Nexus 5"',
        "sec-ch-ua-platform": '"Android"',
        "sec-ch-ua-platform-version": '"6.0"',
        "sec-fetch-dest": "document",
        "sec-fetch-mode": "navigate",
        "sec-fetch-site": "same-origin",
        "sec-fetch-user": "?1",
        "upgrade-insecure-requests": "1",
        "user-agent": "Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/146.0.0.0 Mobile Safari/537.36",
        "cookie": f"cf_clearance={cf_clearance}",
    }

    print(f"[fetch] {shop_id} → {url}")
    try:
        async with httpx.AsyncClient(timeout=30, follow_redirects=True) as client:
            res = await client.get(url, headers=headers)
        print(f"[fetch] {shop_id} status={res.status_code} size={len(res.content)} bytes")
        if res.status_code != 200:
            raise Exception(f"HTTP {res.status_code}: {res.text[:200]}")
        content_type = res.headers.get("content-type", "")
        if "html" in content_type or "json" in content_type:
            raise Exception("Cloudflare chặn — cf_clearance hoặc __cf_chl_tk không hợp lệ")
        content = res.content
    except Exception as e:
        print(f"[fetch] {shop_id} exception: {e}")
        raise

    orders = parse_excel(content, shop_id, shop_name)
    print(f"[parse] {shop_id} → {len(orders)} đơn")

    deleted = db.query(Order).filter(Order.shop_id == shop_id).delete()
    print(f"[delete] {shop_id} → xoá {deleted} đơn cũ")
    for o in orders:
        db.add(Order(**o))
    db.commit()
    unique_order_count = len({str(item.get("order_code", "")).strip() for item in orders if item.get("order_code")})

    meta = db.query(ShopMeta).filter(ShopMeta.shop_id == shop_id).first()
    if meta:
        meta.shop_name   = shop_name
        meta.last_sync   = datetime.now()
        meta.order_count = unique_order_count
    else:
        db.add(ShopMeta(
            shop_id=shop_id, shop_name=shop_name,
            shop_url=shop_url, last_sync=datetime.now(),
            order_count=unique_order_count
        ))
    db.commit()

    print(f"[sync] {shop_name} ({shop_id}): → {len(orders)} đơn")
    return len(orders)
